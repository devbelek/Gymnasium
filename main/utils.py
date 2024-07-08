import csv
from django.http import HttpResponse
import openpyxl

def generate_csv_file(queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students.csv"'

    writer = csv.writer(response)
    writer.writerow(['Имя', 'Фамилия', 'Отчество', 'Класс'])

    for student in queryset:
        writer.writerow([str(student.name), str(student.surname), str(student.last_name), str(student.school_class)])

    return response

def generate_excel_file(queryset):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Students'

    columns = ['Имя', 'Фамилия', 'Отчество', 'Класс']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for student in queryset:
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = str(student.name)
        worksheet.cell(row=row_num, column=2).value = str(student.surname)
        worksheet.cell(row=row_num, column=3).value = str(student.last_name)
        worksheet.cell(row=row_num, column=4).value = str(student.school_class)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    workbook.save(response)

    return response
